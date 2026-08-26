import openpyxl
import pandas as pd
import json
import re
from typing import Optional
from .detector import ParsedDocument, ParsedHolding, clean_num


class ExcelParser:
    def parse(self, file_path: str, file_name: str = "") -> ParsedDocument:
        wb = openpyxl.load_workbook(file_path, data_only=True)

        doc = ParsedDocument()
        doc.account_type = "excel"
        doc.portfolio_name = file_name or file_path
        doc.account_number = re.sub(r'\.xlsx?$', '', file_name, flags=re.IGNORECASE) or "Excel"

        all_rows = []
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            for row in sheet.iter_rows(values_only=True):
                clean_row = [str(cell).strip() if cell is not None else "" for cell in row]
                if any(clean_row):
                    all_rows.append(clean_row)

        if not all_rows:
            return doc

        # Build DataFrame
        max_cols = max(len(row) for row in all_rows)
        df = pd.DataFrame([row + [""] * (max_cols - len(row)) for row in all_rows])

        # Detect header row
        header_idx = -1
        for idx, row in df.iterrows():
            row_str = " ".join([str(val).lower() for val in row])
            if "value" in row_str or "price" in row_str or "rate" in row_str:
                header_idx = idx
                break

        if header_idx != -1:
            headers = [str(h).strip().lower() if str(h).strip() else f"unnamed_{i}"
                       for i, h in enumerate(df.iloc[header_idx].values)]
            df.columns = headers
            df = df.iloc[header_idx + 1:].reset_index(drop=True)

        df = df.loc[:, ~df.columns.str.contains('^unnamed')]

        # Store RAW payloads
        doc.raw_markdown = df.to_markdown(index=False)
        doc.raw_json = json.dumps(df.to_dict(orient='records'))

        # LLM Header Mapping (same approach as PDF)
        from app.llm import call_llm_json
        raw_columns = list(df.columns)
        sys_prompt = """You are a financial data mapping AI.
Given a list of raw bank statement headers, map them to our strict database schema.
Valid Columns: ["security_name", "current_value", "total_cost", "gain_pct", "weight_pct"]

Map each raw header to the closest valid database column.
CRITICAL RULES:
1. "current_value" means the total absolute market value (e.g. "Value", "Amount", "Market Value", "Total Value").
2. "total_cost" means the TOTAL invested amount (e.g. "Cost", "Invested", "Total Cost").
3. NEVER map "Price", "Rate", "NAV", "CMP", "Unit Cost", or "Average Cost" to ANY of the valid columns. Ignore them completely.
4. Ignore irrelevant columns (like "ISIN", "Status", "Account Type", "Scrip Type", "Quantity").
Return ONLY a JSON dictionary: {"raw_header_name": "valid_database_column"}"""

        try:
            mapping = call_llm_json(sys_prompt, str(raw_columns))
            if mapping:
                df = df.rename(columns=mapping)
            else:
                return doc
        except Exception as e:
            print(f"Excel LLM Header Mapping Failed: {e}")
            return doc

        doc.has_cost_data = 'total_cost' in df.columns

        for r in df.to_dict(orient='records'):
            if 'security_name' not in r or 'current_value' not in r:
                continue
            name = str(r['security_name']).strip()
            val = clean_num(r['current_value'])

            if not name or val == 0 or "Total" in name or str(r['current_value']) == 'nan':
                continue

            h = ParsedHolding()
            h.security_name = name.replace(' LTD', '').replace(' LIMITED', '').title()
            h.current_value = val
            h.total_cost = clean_num(r.get('total_cost', 0))
            h.gain_pct = clean_num(r.get('gain_pct', 0))
            h.weight_pct = clean_num(r.get('weight_pct', 0))
            h.asset_class = "Cash and Equivalent" if "cash" in name.lower() or "payable" in name.lower() or "receivable" in name.lower() else "Equity"
            doc.holdings.append(h)

        return doc
